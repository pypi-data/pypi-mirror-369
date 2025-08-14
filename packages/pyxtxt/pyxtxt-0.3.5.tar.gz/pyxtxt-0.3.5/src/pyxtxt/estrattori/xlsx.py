from . import register_extractor
import io
import zipfile
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    import openpyxl
except ImportError:
    openpyxl = None

if openpyxl:

 def xtxt_xlsx(file_buffer, max_rows_per_sheet: int = 200) -> str:
    try:
        file_buffer.seek(0)
        data = file_buffer.read()
        buffer_copy = io.BytesIO(data)

        if not zipfile.is_zipfile(buffer_copy):
            print("⚠️  Invalid XLSX (not a ZIP archive)")
            return ""

        buffer_copy.seek(0)
        wb = openpyxl.load_workbook(buffer_copy, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠️ Error while reading XLSX :  {e}")
        return ""

    testo = []
    for sheet in wb.worksheets:
        if sheet.sheet_state != 'visible':
            continue
        testo.append(f"# {sheet.title}")
        count = 0
        for row in sheet.iter_rows(values_only=True):
            if max_rows_per_sheet != -1 and count >= max_rows_per_sheet:
                break
            valori = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(valori):
                testo.append(" | ".join(valori))
                count += 1

    return "\n".join(testo)

 register_extractor(
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    xtxt_xlsx,
    name="XLSX"
 )


