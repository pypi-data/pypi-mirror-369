"""
AIStructDynSolve
A framework focused on solving structural dynamics problems using artificial intelligence (AI) methods
solve the following ODE of MDOF
M*U_dotdot+C*U_dot+K*U=Pt
initial condition: U(t=0)=InitialU, and U_dot(t=0)=InitialU_dot
Author: 杜轲 duke@iem.ac.cn
Date: 2024/10/26
"""
import torch
import matplotlib.pyplot as plt
import time
import numpy as np
from torch.optim import LBFGS  # L-BFGS优化算法
from ..config import DEVICE

class TrainerInverseProblem:
    def __init__(self, net, loss_functions, Adamsteps, LBFGSsteps, loss_weight=None, HardBC=False):
        if HardBC:
            if loss_weight is None:
                loss_weight = [1.0, 10000.0]
        else:
            if loss_weight is None:
                loss_weight = [1.0, 1.0, 1.0, 10000.0]  # 默认使用四个损失权重

        self.HardBC = HardBC
        self.lossweight = loss_weight
        self.net = net
        self.loss_functions = loss_functions
        self.Adamsteps = Adamsteps
        self.LBFGSsteps = LBFGSsteps
        self.duration_time = loss_functions.gettime()
        self.trainable_params = loss_functions.get_trainable_params()
        self.trainable_params_name = loss_functions.get_trainable_params_name()
        self.dimensionsless_unit = loss_functions.get_dimensionsless_unit()
        self.dimensionsless_mass = loss_functions.get_dimensionsless_mass()
        self.dimensionsless_stiffness = loss_functions.get_dimensionsless_stiffness()
        self.dimensionsless_damping = loss_functions.get_dimensionsless_damping()
        self.dimensionsless_Pt = loss_functions.get_dimensionsless_Pt()
        for name in self.trainable_params_name:
            if name == "pt":
                self.pt_value_scale = loss_functions.get_pt_value_scale()
            elif name == "eq":
                self.eq_acc_scale = loss_functions.get_eq_acc_scale()
            elif name == "eqx":
                self.eq_accX_scale = loss_functions.get_eq_accX_scale()
            elif name == "eqy":
                self.eq_accY_scale = loss_functions.get_eq_accY_scale()
            elif name == "eqz":
                self.eq_accZ_scale = loss_functions.get_eq_accZ_scale()
        self.total_training_time = 0  # 记录总训练时间
        self.losses = []   # 定义损失记录列表
        self.train_params = [] # 定义训练参数记录列表
        self.results = {}

    def train_with_adam(self, init_lr=0.001, scheduler_type='StepLR', step_size=1000, gamma=1.0,T_max=200, factor=0.1, patience=10, show_step=100):
        opt = torch.optim.Adam(params=list(self.net.parameters()) + self.trainable_params, lr=init_lr)
        # 选择学习率调度器
        if scheduler_type == 'StepLR':
            scheduler = torch.optim.lr_scheduler.StepLR(opt, step_size=step_size, gamma=gamma)
        elif scheduler_type == 'ExponentialLR':
            scheduler = torch.optim.lr_scheduler.ExponentialLR(opt, gamma=gamma)
        elif scheduler_type == 'CosineAnnealingLR':
            scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(opt, T_max=T_max)
        elif scheduler_type == 'ReduceLROnPlateau':
            scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(opt, mode='min', factor=factor, patience=patience)
        else:
            raise ValueError(f"Unsupported scheduler type: {scheduler_type}")

        # 使用Adam训练
        print("Start training......")
        print(f"Start Adam training for {self.Adamsteps} steps")
        if self.HardBC:
            print("Step    ", "   Training loss(lossODE,lossObserve)       ", "Training total loss")
        else:
            print("Step    ", "   Training loss(lossODE,lossInitialU,lossInitialU_dot,lossObserve)       ", "Training total loss")

        # 开始计时
        start_time = time.time()
        for i in range(self.Adamsteps):
            opt.zero_grad()
            if self.HardBC:
                totalloss = self.lossweight[0] * self.loss_functions.lossODE() + \
                            self.lossweight[1] * self.loss_functions.lossObserve()
            else:
                totalloss = self.lossweight[0] * self.loss_functions.lossODE() + \
                            self.lossweight[1] * self.loss_functions.lossInitialU() + \
                            self.lossweight[2] * self.loss_functions.lossInitialU_dot() + \
                            self.lossweight[3] * self.loss_functions.lossObserve()
            plt.ion()  # 开启交互模式
            if i % show_step == 0:
                if self.HardBC:
                    print(i, "   ", [self.loss_functions.lossODE().item(), self.loss_functions.lossObserve().item()], "   ",[totalloss.item()])
                else:
                    print(i, "   ", [self.loss_functions.lossODE().item(),
                                         self.loss_functions.lossInitialU().item(),
                                         self.loss_functions.lossInitialU_dot().item(),
                                         self.loss_functions.lossObserve().item()], "   ",
                          [totalloss.item()])
                for name, params_value in zip(self.trainable_params_name, self.trainable_params):
                    if name == "mass":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_mass
                    elif name == "stiffness":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_stiffness
                    elif name == "damping":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_damping
                    elif name == "pt":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_Pt / self.pt_value_scale
                    elif name == "eq":
                        params = params_value * self.dimensionsless_Pt / self.eq_acc_scale
                    elif name == "eqx":
                        params = params_value * self.dimensionsless_Pt / self.eq_accX_scale
                    elif name == "eqy":
                        params = params_value * self.dimensionsless_Pt / self.eq_accY_scale
                    elif name == "eqz":
                        params = params_value * self.dimensionsless_Pt / self.eq_accZ_scale
                    else:
                        params = params_value
                    # 使用 f-string 进行格式化输出
                    print(f"Updated {name}: {params.cpu().detach().numpy().copy()}")
                t_test = torch.linspace(0, self.duration_time, 2000, device=DEVICE).unsqueeze(1)  # 设置测试时间
                with torch.no_grad():
                    u_pred = self.net(t_test)* self.dimensionsless_Pt   # 计算预测位移
                    plt.cla()
                    for j in range(u_pred.shape[1]):  # 遍历每组数据
                        plt.plot(t_test.cpu().numpy(), u_pred.detach().cpu()[:, j], label=f'DOF {j + 1}', linewidth=3)
                    plt.legend()
                    plt.xlabel('t_test')
                    plt.ylabel('u_pred')
                    plt.title('Predicted Displacement for All DOFs')
                    plt.show()
                    plt.pause(0.1)
            totalloss.backward()
            opt.step()
            scheduler.step()
            self.losses.append(totalloss.item())
            for name,params_value in zip(self.trainable_params_name, self.trainable_params):
                if name == "mass":
                    params_adam = params_value * self.dimensionsless_unit * self.dimensionsless_mass
                elif name == "stiffness":
                    params_adam = params_value * self.dimensionsless_unit * self.dimensionsless_stiffness
                elif name == "damping":
                    params_adam = params_value * self.dimensionsless_unit * self.dimensionsless_damping
                elif name == "pt":
                    params_adam = params_value * self.dimensionsless_unit * self.dimensionsless_Pt / self.pt_value_scale
                elif name == "eq":
                    params_adam = params_value * self.dimensionsless_Pt / self.eq_acc_scale
                elif name == "eqx":
                    params_adam = params_value * self.dimensionsless_Pt / self.eq_accX_scale
                elif name == "eqy":
                    params_adam = params_value * self.dimensionsless_Pt / self.eq_accY_scale
                elif name == "eqz":
                    params_adam = params_value * self.dimensionsless_Pt / self.eq_accZ_scale
                else:
                    params_adam = params_value
                # 将转换后的结果添加到 train_params 列表中
                self.train_params.append({
                    name: params_adam.cpu().detach().numpy().copy()  # 将张量转移到 CPU，detach 并转换为 NumPy 数组
                })

        end_time = time.time()
        adam_training_time = end_time - start_time
        self.total_training_time += adam_training_time  # 更新总训练时间
    def train_with_lbfgs(self, max_iter=1000, tolerance_grad=0, tolerance_change=0, line_search_fn='strong_wolfe',show_step=100):
        lbfgs_opt = LBFGS(params=list(self.net.parameters()) + self.trainable_params, max_iter=max_iter, tolerance_grad=tolerance_grad, tolerance_change=tolerance_change, line_search_fn=line_search_fn)
        lbfgsiteration = 0
        start_time = time.time()

        def closure():
            nonlocal lbfgsiteration
            lbfgs_opt.zero_grad()
            if self.HardBC:
                totalloss = self.lossweight[0] * self.loss_functions.lossODE() + \
                            self.lossweight[1] * self.loss_functions.lossObserve()
            else:
                totalloss = self.lossweight[0] * self.loss_functions.lossODE() + \
                            self.lossweight[1] * self.loss_functions.lossInitialU() + \
                            self.lossweight[2] * self.loss_functions.lossInitialU_dot() + \
                            self.lossweight[3] * self.loss_functions.lossObserve()
            if lbfgsiteration % show_step == 0:
                if self.HardBC:
                    print(lbfgsiteration, "   ", [self.loss_functions.lossODE().item(), self.loss_functions.lossObserve().item()], "   ",[totalloss.item()])
                else:
                    print(lbfgsiteration, "   ", [self.loss_functions.lossODE().item(),
                                         self.loss_functions.lossInitialU().item(),
                                         self.loss_functions.lossInitialU_dot().item(),
                                         self.loss_functions.lossObserve().item()], "   ",
                          [totalloss.item()])
                for name, params_value in zip(self.trainable_params_name, self.trainable_params):
                    # 使用 f-string 进行格式化输出
                    if name == "mass":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_mass
                    elif name == "stiffness":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_stiffness
                    elif name == "damping":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_damping
                    elif name == "pt":
                        params = params_value * self.dimensionsless_unit * self.dimensionsless_Pt / self.pt_value_scale
                    elif name == "eq":
                        params = params_value * self.dimensionsless_Pt / self.eq_acc_scale
                    elif name == "eqx":
                        params = params_value * self.dimensionsless_Pt / self.eq_accX_scale
                    elif name == "eqy":
                        params = params_value * self.dimensionsless_Pt / self.eq_accY_scale
                    elif name == "eqz":
                        params = params_value * self.dimensionsless_Pt / self.eq_accZ_scale
                    else:
                        params = params_value
                    print(f"Updated {name}: {params.cpu().detach().numpy().copy()}")
                t_test = torch.linspace(0, self.duration_time, 2000, device=DEVICE).unsqueeze(1)
                with torch.no_grad():
                    u_pred = self.net(t_test)* self.dimensionsless_Pt
                    plt.cla()
                    for i in range(u_pred.shape[1]):
                        plt.plot(t_test.cpu().numpy(), u_pred.cpu().numpy()[:, i], label=f'DOF {i + 1}', linewidth=3)
                    plt.legend()
                    plt.xlabel('t_test')
                    plt.ylabel('u_pred')
                    plt.title('Predicted Displacement for All DOFs')
                    plt.show()
                    plt.pause(0.1)
            totalloss.backward()
            self.losses.append(totalloss.item())
            for name,params_value in zip(self.trainable_params_name, self.trainable_params):
                if name == "mass":
                    params_lbfgs = params_value * self.dimensionsless_unit * self.dimensionsless_mass
                elif name == "stiffness":
                    params_lbfgs = params_value * self.dimensionsless_unit * self.dimensionsless_stiffness
                elif name == "damping":
                    params_lbfgs = params_value * self.dimensionsless_unit * self.dimensionsless_damping
                elif name == "pt":
                    params_lbfgs = params_value * self.dimensionsless_unit * self.dimensionsless_Pt / self.pt_value_scale
                elif name == "eq":
                    params_lbfgs = params_value * self.dimensionsless_Pt / self.eq_acc_scale
                elif name == "eqx":
                    params_lbfgs = params_value * self.dimensionsless_Pt / self.eq_accX_scale
                elif name == "eqy":
                    params_lbfgs = params_value * self.dimensionsless_Pt / self.eq_accY_scale
                elif name == "eqz":
                    params_lbfgs = params_value * self.dimensionsless_Pt / self.eq_accZ_scale
                else:
                    params_lbfgs = params_value
                # 将转换后的结果添加到 train_params 列表中
                self.train_params.append({
                    name: params_lbfgs.cpu().detach().numpy().copy()  # 将张量转移到 CPU，detach 并转换为 NumPy 数组
                })

            lbfgsiteration += 1  # 更新迭代次数

            if lbfgsiteration >= self.LBFGSsteps:  # 检查是否达到最大迭代次数
                raise StopIteration  # 抛出异常以停止优化

            return totalloss

        print(f"Start L-BFGS training for {self.LBFGSsteps} steps")
        try:
            while True:
                lbfgs_opt.step(closure)
        except StopIteration:
            print("Training finished")
            # 结束计时
            end_time = time.time()
            lbfgs_training_time = end_time - start_time
            self.total_training_time += lbfgs_training_time  # 更新总训练时间
            print(f"Total training time: {self.total_training_time} seconds")
            # 保存训练结果，将模型和标量存储在一个字典中
            self.results = {
                'net': self.net,
                'dimensionsless_Pt': self.dimensionsless_Pt,
                'duration_time': self.duration_time,
            }

            # 保存更新参数值
            # 打开文件写入数据
            trainable_params_name = self.trainable_params_name  # 比如 ["mass", "stiffness", "damping"]
            params_list = self.train_params  # 一个列表，每个元素是一个只含一个参数的字典

            group_size = len(trainable_params_name)
            merged_params = []

            # 按 group_size 分组合并字典
            for i in range(0, len(params_list), group_size):
                combined = {}
                for j in range(group_size):
                    idx = i + j
                    if idx < len(params_list):
                        combined.update(params_list[idx])
                merged_params.append(combined)

            # 写文件
            with open('train_params_matrix.txt', 'w') as f:
                for idx, param_group in enumerate(merged_params):
                    f.write(f"=== Training Step {idx + 1} ===\n")
                    for name in trainable_params_name:
                        f.write(f"{name}:\n")
                        matrix = param_group.get(name)
                        if matrix is not None:
                            if matrix.ndim == 0:
                                # 标量直接格式化为浮点数，不用 array2string
                                f.write(f"{matrix.item():.8f}\n")
                            else:
                                # 1D 或 2D 都用 array2string 处理
                                matrix_str = np.array2string(
                                    matrix,
                                    separator=', ',
                                    threshold=1000000,
                                    max_line_width=1000000,
                                    formatter={'all': lambda x: f'{x:.8f}'}
                                )
                                f.write(matrix_str + '\n')
                        else:
                            f.write("None\n")
                    f.write('\n')

            print("Data saved to train_params_matrix.txt")
            from openpyxl import Workbook
            # 2. 创建 Excel
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认 sheet

            # 3. 为每个参数建一个 sheet
            for param_name in trainable_params_name:
                ws = wb.create_sheet(title=param_name)

                row_cursor = 1
                for idx, param_group in enumerate(merged_params):
                    matrix = param_group.get(param_name)

                    # 写 PARAM SET 名
                    ws.cell(row=row_cursor, column=1, value=f"Training Step {idx + 1}")

                    if matrix is not None:
                        if matrix.ndim == 0:
                            # 标量 → 写入单个值
                            ws.cell(row=row_cursor, column=2, value=float(matrix))
                            row_cursor += 1
                        elif matrix.ndim == 1:
                            # 1D 向量 → 写一行
                            for j in range(matrix.shape[0]):
                                ws.cell(row=row_cursor, column=2 + j, value=float(matrix[j]))
                            row_cursor += 1
                        elif matrix.ndim == 2:
                            matrix = np.atleast_2d(matrix)  # 保证是 2D
                            rows, cols = matrix.shape
                            for i in range(rows):
                                for j in range(cols):
                                    value = float(matrix[i, j])
                                    ws.cell(row=row_cursor + i, column=2 + j, value=value)
                            row_cursor += rows   # 加空行
                        else:
                            raise ValueError(f"Unsupported matrix dimension: {matrix.ndim}")
                    else:
                        ws.cell(row=row_cursor, column=2, value="None")
                        row_cursor += 1  # 多加一行分隔

            # 4. 保存 Excel 文件
            wb.save("train_params_by_parameter.xlsx")
            print("Saved as train_params_by_parameter.xlsx")


