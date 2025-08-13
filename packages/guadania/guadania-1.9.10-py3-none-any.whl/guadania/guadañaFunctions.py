from .prisma.prismaFunctions import PrismaSession
from . import prisma

import os
import datetime
import json
import pandas as pd

from openpyxl import load_workbook


def json_to_df(json):
    return pd.json_normalize(json)


def df_to_jsons_list(df):
    """
    The opposite of json_normalize
    """
    result = []
    for idx, row in df.iterrows():
        parsed_row = {}
        for col_label, v in row.items():
            keys = col_label.split(".")

            current = parsed_row
            for i, k in enumerate(keys):
                if i == len(keys) - 1:
                    current[k] = v
                else:
                    if k not in current.keys():
                        current[k] = {}
                    current = current[k]
        # save
        result.append(parsed_row)
    return result


def to_excel(dst, table, append=False):
    """
    Exporta a una tabla de Excel. Puede escribir al final de una ya creada también.

    :param dst: Ruta de destino. Debe contener el nombre del archivo en ella. P.e.: "./excel.xlsx"
    :param table: DataFrame a exportar a Excel
    :param append: Si True, añade al final de la hoja existente.
    """

    if append and os.path.exists(dst):
        wb = load_workbook(dst)
        sheet = wb.worksheets[0]
        startrow = sheet.max_row

        with pd.ExcelWriter(dst, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            table.to_excel(writer, index=False, header=False, startrow=startrow)

    else:
        table.to_excel(dst, engine='openpyxl', index=False)


def cloud_accounts(prismasession):
    """
    Obtiene un DataFrame con todas las cuentas del tenant.

    :param prismasession: datos de la sesión con el tenant (url, token)
    :return: DataFrame con cada una de las cuentas del tenant
    """

    res = prisma.list_cloud_accounts(prismasession)
    df = pd.json_normalize(res)
    orgs = df[df['numberOfChildAccounts'] > 0]
    pd.set_option('display.max_columns', None)

    for org, cloud in zip(list(orgs.accountId), list(orgs.cloudType)):
        childs = cloud_org_accounts(prismasession, cloud, org)
        df = pd.concat([df, childs])

    return df


def cloud_org_accounts(prismasession, cloud_type, id, return_json=False):
    """
    Obtiene un DataFrame con todas las cuentas hijas de una organización.

    :param prismasession: datos de la sesión con el tenant (url, token)
    :param cloud_type: tipo de nube
    :param id: ID de la cuenta organización
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con cada una de las cuentas hijas de una organización
    """

    _cloud_type = cloud_type.lower()
    cloud_list = ['aws', 'azure', 'gcp', 'oci']
    if (_cloud_type in cloud_list):
        res = prisma.list_cloud_org_accounts(prismasession, _cloud_type, id)

        if (return_json):
            return res
        else:
            return pd.json_normalize(res)
    else:
        raise Exception('Error introduciendo tipo de nube. Valores aceptables: aws, azure, gcp, oci.')


def account_groups(prismasession, detailed=True, return_json=False):
    """
    Obtiene un DataFrame con todas las account groups del tenant.

    :param detailed: True si queremos info detallada.
    :param prismasession: datos de la sesión con el tenant (url, token).
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con cada una de las account group del tenant.
    """

    res = prisma.list_account_groups(prismasession, detailed)
    if (return_json):
        return res
    return pd.json_normalize(res)


def account_groups_by_cloud_type(prismasession, cloud_type=None, return_json=False):
    """
    Obtiene un DataFrame con todas las account groups del tenant por tipo de nube.

    :param cloud_type: definir el tipo de cloud [aws,gcp,azure].
    :param prismasession: datos de la sesión con el tenant (url, token).
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con cada una de las account group del tenant.
    """

    res = prisma.list_account_groups_by_cloud_type(prismasession, cloud_type)
    if (return_json):
        return res
    return pd.json_normalize(res)


def compliance_summary(prismasession: PrismaSession, standards, return_json=False):
    """
    Obtiene un DataFrame con el resumen de compliance de un compliance standard de Prisma.

    Obtiene: failedResources, passedResources, totalResources, highSeverityFailedResources,
    mediumSeverityFailedResources y lowSeverityFailedResources.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param complianceStandard: Standards de los que se busca el compliance (una lista).
    :param return_json: Devuelve la información en formato JSON

    :return: DataFrame con el resumen compliance.
    """

    tenantStandards = compliance_standards(prismasession)
    summaries = []
    for standard in standards:
        selected_id = tenantStandards[tenantStandards['name'] == standard]['id'].values[0]
        res = prisma.get_standard_compliance_statistics(prismasession, selected_id)
        summaries.append(res)

    if (return_json):
        return summaries
    else:
        return pd.json_normalize(summaries)


def compliance_requirements_summary(prismasession: PrismaSession, standard, account_name=None, return_json=False):
    """
    Obtiene un DataFrame con el resumen de compliance de los requerimientos de un compliance standard de Prisma.

    Obtiene: failedResources, passedResources, totalResources, highSeverityFailedResources,
    mediumSeverityFailedResources y lowSeverityFailedResources.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param standard: Standard de los que se busca los requerimientos para compliance.
    :param return_json: Devuelve la información en formato JSON

    :return: DataFrame con el resumen compliance.
    """

    tenantStandards = compliance_standards(prismasession)
    standard_id = tenantStandards[tenantStandards['name'] == standard]['id'].values[0]

    comp = {}
    requirements = standard_requirements(prismasession, standard)
    print(requirements)
    for r in list(requirements['id']):
        reqcomp = prisma.get_standard_requirement_compliance_statistics(prismasession, standard_id, r, account_name)
        comp[requirements.loc[requirements['id'] == r, 'requirementId'].iloc[0]] = reqcomp

    if (return_json):
        return comp
    else:
        dfs = []
        for req in comp.keys():
            df = pd.json_normalize(comp[req])
            df['RequirementID'] = req
            dfs.append(df.copy())

        df = pd.concat(dfs, axis=0)
        print(df)
        return df


def compliance_section_summary(prismasession, complianceStandard, accountName, return_json=False):
    """
    Obtiene un DataFrame con el compliance de cada seccion de un compliance standard de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param complianceStandard: compliance standard del que se busca los resultados
    :param accountName: Nombre de la cuenta.
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con compliance detallado por resource de un compliance standard de Prisma
    """

    standards = compliance_standards(prismasession)
    selected_name = standards.loc[standards['name'] == complianceStandard].head(1).iloc[0]['name']
    selected_id = standards.loc[standards['name'] == complianceStandard].head(1).iloc[0]['id']

    requirements = standard_requirements(prismasession, selected_name)
    print(requirements)
    comp = {}
    for r in list(requirements['id']):
        reqcomp = prisma.get_standard_requirement_compliance_statistics(prismasession, selected_id, r, accountName)
        comp[requirements.loc[requirements['id'] == r, 'requirementId'].iloc[0]] = reqcomp

    if (return_json):
        return comp
    else:
        dfs = []
        for req in comp.keys():
            df = pd.json_normalize(comp[req]['complianceDetails'])
            df['RequirementID'] = req
            dfs.append(df.copy())

        df = pd.concat(dfs, axis=0)
        print(df)
        return df

    res = []
    for summary in reqcomp["complianceDetails"]:
        res.append(summary)

    if (return_json):
        return res
    return pd.json_normalize(res)


def filter_and_unroll(lst, column, target):
    filtered_dicts = [item for item in lst if item.get(column) == target]
    if filtered_dicts:
        return filtered_dicts
    else:
        return []


# Si metes el account id no mete el account group
def compliance_section_by_policy(prismasession, complianceStandard, accountName=None, accountGroup=None,
                                 return_json=False):
    section_summary = compliance_section_summary(prismasession, complianceStandard, accountName, return_json=True)

    all_alerts = alerts(prismasession, policyComplianceStandard=complianceStandard, basic=False, detailed=True,
                        cloudAccount=accountName, accountGroup=accountGroup)

    if not all_alerts.empty:
        section_alerts = all_alerts[['id', 'resource.id', 'policy.complianceMetadata']]
    else:
        section_alerts = all_alerts
    failed_resources = {}
    policies = list_policies(prismasession, complianceStandard)
    section_req = {}
    for requirement, compDetails in section_summary.items():
        if compDetails['summary']['totalResources'] == 0:
            continue
        for section in compDetails['complianceDetails']:
            if section['totalResources'] == 0:
                continue
            else:
                failed_resources[section['name']] = {}
                for index, pol in policies.iterrows():
                    for comp in pol['complianceMetadata']:
                        if (comp['standardName'] == complianceStandard):
                            # print(section['complianceDetails'][0]['name'])
                            if (requirement == comp['requirementId'] and section['name'] == comp['sectionId']):
                                section_req[section['name']] = comp["requirementId"]
                                failed_resources[section['name']][pol['name']] = {'failed': 0,
                                                                                  'total': section['totalResources'],
                                                                                  'severity': pol['severity']}

        for index, row in section_alerts.iterrows():
            pol_comp = row['policy.complianceMetadata']
            for comp in pol_comp:
                if (comp['standardName'] == complianceStandard and requirement == comp['requirementId'] and comp[
                    'sectionId'] == section['name']):
                    polname = policies[policies['policyId'] == comp['policyId']]['name'].item()
                    failed_resources[section['name']][polname]['failed'] += 1

    df = {"accountGroup": [], "subscriptionId": [], "standard": [], "requirement": [], "section": [], "policy": [],
          "severity": [], "failed": [], "totalSection": []}
    for sec in failed_resources.keys():
        for pol in failed_resources[sec].keys():
            df["accountGroup"].append(accountGroup)
            df["subscriptionId"].append(accountName)
            df["standard"].append(complianceStandard)
            df["requirement"].append(section_req[sec])
            df["section"].append(sec)
            df["policy"].append(pol)
            df["severity"].append(failed_resources[sec][pol]["severity"])
            df["failed"].append(failed_resources[sec][pol]["failed"])
            df["totalSection"].append(failed_resources[sec][pol]["total"])

    todel = []
    for sec in failed_resources.keys():
        if failed_resources[sec] == {}:
            todel.append(sec)

    for sec in todel:
        failed_resources.pop(sec)

    if return_json:
        return failed_resources
    else:
        df = pd.DataFrame(df)

        return df


def historic(prismasession: PrismaSession, complianceStandard, cloudAccounts=None, accountGroups=None):
    """
    Obtiene un DataFrame con el historial de compliance de un compliance standard de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param complianceStandard: compliance standard del que se busca el compliance
    :param cloudAccounts: lista opcional con los id de los cloud accounts en caso de que se quiera separar el historico por cuentas. Sólo aplica si no se define 'accountGroups'.
    :param accountGroups: lista opcional con los names de los accountGroups en caso de que se quiera separar el historico por cuentas. Tiene preferencia sobre 'cloudAccounts'.
    :return: DataFrame con el historial de compliance de un compliance standard de Prisma
    """

    if accountGroups:
        res = pd.DataFrame()
        groupIDs = get_group_accounts_ids_by_names(prismasession, accountGroups)
        cloudAccountsDetails = prisma.list_cloud_accounts_names(prismasession, accountGroupIds=groupIDs)

        cloudAccountsNames = [account['name'] for account in cloudAccountsDetails]

        return historic(prismasession, complianceStandard, cloudAccountsNames)

    standards = compliance_standards(prismasession)
    selected_id = standards.loc[standards['name'] == complianceStandard].head(1).iloc[0]['id']

    if cloudAccounts:
        res = pd.DataFrame()
        for account in cloudAccounts:
            tmp = pd.json_normalize(
                prisma.get_standard_compliance_statistics_trend(prismasession, selected_id, account))
            tmp['AccountName'] = account
            res = pd.concat([res, tmp])

    else:
        res = pd.json_normalize(prisma.get_standard_compliance_statistics_trend(prismasession, selected_id))

    res['date'] = res.apply(lambda x: timestamp_to_date(x), 1)

    return res


def get_group_accounts_ids_by_names(prismaSession: PrismaSession, groupAccountNames):
    ids = []
    res = prisma.list_account_groups(prismaSession)
    for group in res:
        if group['name'] in groupAccountNames:
            ids.append(group['id'])

    return ids


def timestamp_to_date(x):
    _x = json.loads(x.to_json())
    date = datetime.datetime.fromtimestamp(_x['timestamp'] / 1000).strftime("%d/%m/%Y")
    return date


def compliance_standards(prismasession, return_json=False):
    """
    Obtiene un DataFrame con todos los Standards (compliance standards) de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con cada uno de los standards de Prisma
    """

    res = prisma.list_compliance_standards(prismasession)
    if (return_json):
        return res
    return pd.json_normalize(res)


def standard_requirements(prismasession, complianceStandard, return_json=False):
    """
    Obtiene todos los requerimientos de un compliance standard de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param complianceStandard: nombre del compliance standard
    :return: DataFrame con cada uno de los standards de Prisma
    """

    standards = compliance_standards(prismasession)
    try:
        selected_id = standards.loc[standards['name'] == complianceStandard].head(1).iloc[0]['id']
    except:
        raise Exception(
            "Error obteniendo requiermientos. Probablemente, nombre de complianceStandard incorrecto.")

    res = prisma.list_compliance_requirements(prismasession, selected_id)
    if (return_json):
        return res
    return pd.json_normalize(res)


def standards_sections(prismasession, complianceStandard, requirement_name, return_json=False):
    """
    Obtiene todas las secciones de un compliance standard de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param compliance standard: nombre del compliance standard
    :param requirement_name: nombre del requerimiento
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con cada uno de las secciones de un requerimiento de Prisma
    """

    reqs = standard_requirements(prismasession, complianceStandard)
    try:
        req_id = reqs.loc[reqs['name'] == requirement_name].head(1).iloc[0]['id']
    except:
        raise Exception("Error obteniendo requiermientos. Probablemente, nombre de compliance standard incorrecto.")

    res = prisma.list_compliance_sections(prismasession, req_id)
    if (return_json):
        return res
    return pd.json_normalize(res)


def resources_usage_over_time(prismaSession, accounts_ids, start_date=None,
                              end_date=None, return_json=False):
    """
    Obtiene un json con el uso de recursos licenciables sobre el tiempo. Si no se concreta fecha de inicio o fecha de final,
    se asume que se busca sin acotación de fechas. Para acotar, hay que definir ambas fechas, inicio y final.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param accounts_ids: Lista con IDs de las suscripciones sobre las que obtener el uso.
    :param start_date: Fecha de inicio, en formato "dd/mm/yyyy". Por defecto, es el origen.
    :param end_date: Fecha de fin, en formato "dd/mm/yyyy". Por defecto, es hoy.
    :param return_json: Devuelve la información en formato JSON
    :return: DataFrame con el uso de recursos licenciables sobre el tiempo.
    """

    time_range = None
    if (start_date != None and end_date != None):
        time_range = {
            "type": "absolute",
            "value": {
                "startTime": int(datetime.datetime.strptime(start_date, '%d/%m/%Y').timestamp() * 1000),
                "endTime": int(datetime.datetime.strptime(end_date, '%d/%m/%Y').timestamp() * 1000)
            }
        }

    res = prisma.resource_usage_over_time(
        prismaSession, accounts_ids, time_range)
    if (return_json):
        return res
    return pd.json_normalize(res)


def list_policies(prismaSession, complianceStandard):
    """
    Obtiene un DataFrame con un listado de politicas para un Compliance standard.

    La API de prisma no permite listar políticas pasando cierto standard. En cambio,
    al listar políticas, en sus detalles aparece el standard al que pertenecen. Por lo tanto,
    se listan las politicas y comprobamos si el standard aparece en sus detalles. Una policy
    puede pertenecer a varios standard.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param complianceStandard: nombre del compliance standard de las politicas a buscar
    :return: DataFrame con un listado de politicas.
    """

    res_json = prisma.list_policies(prismaSession)
    res = pd.json_normalize(res_json)
    res['compliance-standard'] = res.apply(lambda x: list_policies_get_standard(x, complianceStandard), 1)
    res = res.loc[res['compliance-standard'] == complianceStandard]
    return res


# Con esta función podemos comprobar si la policy pertenece a un determinado compliance standard
def list_policies_get_standard(x, complianceStandard):
    _x = json.loads(x.to_json())

    if ('complianceMetadata' in _x.keys() and _x['complianceMetadata'] != None):
        for c in _x['complianceMetadata']:
            if (c['standardName'] == complianceStandard):
                return complianceStandard
        return ''
    else:
        return ''


def alerts(prismasession: prisma.PrismaSession, basic=True, timeType='relative', timeAmount='5',
           timeUnit: prisma.TimeUnit = prisma.TimeUnit.YEAR, detailed=False, sortBy=None, offset=0,
           alertId=None, alertStatus: prisma.AlertStatus = prisma.AlertStatus.OPEN, cloudAccount=None,
           cloudAccountId=None, accountGroup=None, cloudType=None, cloudRegion=None,
           cloudService=None, policyId=None, policyName=None, policySeverity: prisma.PolicySeverity = None,
           policyLabel=None, policyType: prisma.PolicyType = None, policyComplianceStandard=None,
           policyComplianceRequirement=None, policyComplianceSection=None, policyRemediable=None,
           alertRuleName=None, resourceId=None, resourceName=None, resourceType=None, tagsAsString=False):
    """
    Obtiene un DataFrame de alertas de Prisma paginada.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param basic: Si queremos sólo los datos de alertas más básicos: id, status, alertTime, resource.resourceType, resource.cloudType, resource.account y policy.policyId.
        Por defecto está a True.
    :param detailed: True si queremos detalles de las alertas, False en caso contrario
    :param timeType: Tipo de tiempo, default 'relative'.
    :param timeAmount: Cantidad de tiempo, unidad definida en el parámetro timeUnit, default 30.
    :param timeUnit: Unidad de tiempo, default 'year'.
    :param detailed: Alerta detallada o no, default 'True'.
    :param sortBy: Ordenar las alertas. El formato del parámetro es PROPERTY:asc para ascendiente, PROPERTY:desc para descendiente. P.e.: sortBy=id:desc, sortBy=firstseen:asc, sortBy=lastseen:desc).
        Propiedades válidas: firstseen, lastseen, resource.regionid, alerttime, id, resource.accountid, status y resource.id.
    :param offset: Número de alertas que saltar (ignorar) en los resultados.
    :param limit: Número máximo de alertas, no más de 10000. Default 10000.
    :param pageToken: Identificador de la página de alertas, para cuando las alertas no caben en una respuesta.
    :param alertId: Alert ID.
    :param alertStatus: Enum AlertStatus.
    :param cloudAccount: Cloud account.
    :param cloudAccountId: ID de la cloud account.
    :param accountGroup: Account group.
    :param cloudType: Cloud type, enum CloudType.
    :param cloudRegion: Cloud region.
    :param cloudService: Cloud service.
    :param policyId: ID de la policy.
    :param policyName: Nombre de la policy.
    :param policySeverity: PolicySeverity enum.
    :param policyLabel: Label de la policy.
    :param policyType: PolicyType enum.
    :param policyComplianceRequirement: Nombre del Compliance Requirement.
    :param policyComplianceStandard: Si queremos filtrar por standard. También trae datos de las políticas asociadas a esas alertas.
    :param policyComplianceSection: ID del Compliance Section.
    :param policyRemediable: Bool. True es remediable, False no.
    :param alertRuleName: Nombre de la Alert Rule.
    :param resourceId: ID del resource.
    :param resourceName: Nombre del resource.
    :param resourceType: Tipo del resource.
    :param tagsAsString: Si es True, devuelve los tags como un string.
    :return: DataFrame con las alertas del tenant
    """

    # Prisma devuelve muchísimos metadatos, vamos a quedarnos con los interesantes
    columns_base = ['id', 'status', 'alertTime', 'resource.resourceType', 'resource.cloudType', 'resource.account',
                    'policy.policyId']

    alerts = []
    nextPageToken = None
    looping = True
    while (looping):
        res = prisma.list_alerts_v2(prismasession, timeType=timeType, timeAmount=timeAmount, timeUnit=timeUnit,
                                    detailed=detailed, sortBy=sortBy,
                                    offset=offset, limit=10000, pageToken=nextPageToken, alertId=alertId,
                                    alertStatus=alertStatus,
                                    cloudAccount=cloudAccount, cloudAccountId=cloudAccountId, accountGroup=accountGroup,
                                    cloudType=cloudType,
                                    cloudRegion=cloudRegion, cloudService=cloudService, policyId=policyId,
                                    policyName=policyName,
                                    policySeverity=policySeverity, policyLabel=policyLabel, policyType=policyType,
                                    policyComplianceStandard=policyComplianceStandard,
                                    policyComplianceRequirement=policyComplianceRequirement,
                                    policyComplianceSection=policyComplianceSection,
                                    policyRemediable=policyRemediable, alertRuleName=alertRuleName,
                                    resourceId=resourceId, resourceName=resourceName,
                                    resourceType=resourceType)

        alerts += res['items']
        if ('nextPageToken' not in res.keys()):
            looping = False
        else:
            nextPageToken = res['nextPageToken']

    if (tagsAsString):
        columns_base.append("tags_as_string")
        for i in range(len(alerts)):
            tags = ""
            try:
                tags = alerts[i]["resource"]["data"]["tags"]
                tags = str(tags)


            except:
                pass

            alerts[i]["tags_as_string"] = tags

    df = pd.json_normalize(alerts)
    if (not df.empty and basic):
        df = df[columns_base]

    if (not df.empty and policyComplianceStandard is not None):
        policies = list_policies(prismaSession=prismasession, complianceStandard=policyComplianceStandard)
        policies = policies[['policyId', 'name', 'description', 'severity', 'recommendation', 'compliance-standard']]
        df = pd.merge(df, policies, how='left', left_on=['policy.policyId'], right_on=['policyId'])
    return df

def alerts_post(prismasession: prisma.PrismaSession, basic=True, timeType='relative', timeAmount='10',
           timeUnit: prisma.TimeUnit = prisma.TimeUnit.YEAR, detailed=False, sortBy=None, offset=0,
           alertId=None, alertStatus: prisma.AlertStatus = prisma.AlertStatus.OPEN, cloudAccount=None,
           cloudAccountId=None, accountGroup=None, cloudType=None, cloudRegion=None,
           cloudService=None, policyId=None, policyName=None, policySeverity: prisma.PolicySeverity = None,
           policyLabel=None, policyType: prisma.PolicyType = None, policyComplianceStandard=None,
           policyComplianceRequirement=None, policyComplianceSection=None, policyRemediable=None,
           alertRuleName=None, resourceId=None, resourceName=None, resourceType=None, tagsAsString=False):
    """
    Obtiene un DataFrame de alertas de Prisma paginada.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param basic: Si queremos sólo los datos de alertas más básicos: id, status, alertTime, resource.resourceType, resource.cloudType, resource.account y policy.policyId.
        Por defecto está a True.
    :param detailed: True si queremos detalles de las alertas, False en caso contrario
    :param timeType: Tipo de tiempo, default 'relative'.
    :param timeAmount: Cantidad de tiempo, unidad definida en el parámetro timeUnit, default 30.
    :param timeUnit: Unidad de tiempo, default 'year'.
    :param detailed: Alerta detallada o no, default 'True'.
    :param sortBy: Ordenar las alertas. El formato del parámetro es PROPERTY:asc para ascendiente, PROPERTY:desc para descendiente. P.e.: sortBy=id:desc, sortBy=firstseen:asc, sortBy=lastseen:desc).
        Propiedades válidas: firstseen, lastseen, resource.regionid, alerttime, id, resource.accountid, status y resource.id.
    :param offset: Número de alertas que saltar (ignorar) en los resultados.
    :param limit: Número máximo de alertas, no más de 10000. Default 10000.
    :param pageToken: Identificador de la página de alertas, para cuando las alertas no caben en una respuesta.
    :param alertId: Alert ID.
    :param alertStatus: Enum AlertStatus.
    :param cloudAccount: Cloud account.
    :param cloudAccountId: ID de la cloud account.
    :param accountGroup: Account group.
    :param cloudType: Cloud type, enum CloudType.
    :param cloudRegion: Cloud region.
    :param cloudService: Cloud service.
    :param policyId: ID de la policy.
    :param policyName: Nombre de la policy.
    :param policySeverity: PolicySeverity enum.
    :param policyLabel: Label de la policy.
    :param policyType: PolicyType enum.
    :param policyComplianceRequirement: Nombre del Compliance Requirement.
    :param policyComplianceStandard: Si queremos filtrar por standard. También trae datos de las políticas asociadas a esas alertas.
    :param policyComplianceSection: ID del Compliance Section.
    :param policyRemediable: Bool. True es remediable, False no.
    :param alertRuleName: Nombre de la Alert Rule.
    :param resourceId: ID del resource.
    :param resourceName: Nombre del resource.
    :param resourceType: Tipo del resource.
    :param tagsAsString: Si es True, devuelve los tags como un string.
    :return: DataFrame con las alertas del tenant
    """

    # Prisma devuelve muchísimos metadatos, vamos a quedarnos con los interesantes
    columns_base = ['id', 'status', 'alertTime', 'resource.resourceType', 'resource.cloudType', 'resource.account',
                    'policy.policyId']

    alerts = []
    nextPageToken = None
    looping = True
    while (looping):
        res = prisma.list_alerts_v2_post(prismasession, timeType=timeType, timeAmount=timeAmount, timeUnit=timeUnit,
                                    detailed=detailed, sortBy=sortBy,
                                    offset=offset, limit=10000, pageToken=nextPageToken, alertId=alertId,
                                    alertStatus=alertStatus,
                                    cloudAccount=cloudAccount, cloudAccountId=cloudAccountId, accountGroup=accountGroup,
                                    cloudType=cloudType,
                                    cloudRegion=cloudRegion, cloudService=cloudService, policyId=policyId,
                                    policyName=policyName,
                                    policySeverity=policySeverity, policyLabel=policyLabel, policyType=policyType,
                                    policyComplianceStandard=policyComplianceStandard,
                                    policyComplianceRequirement=policyComplianceRequirement,
                                    policyComplianceSection=policyComplianceSection,
                                    policyRemediable=policyRemediable, alertRuleName=alertRuleName,
                                    resourceId=resourceId, resourceName=resourceName,
                                    resourceType=resourceType)

        alerts += res['items']
        if ('nextPageToken' not in res.keys()):
            looping = False
        else:
            nextPageToken = res['nextPageToken']

    if (tagsAsString):
        columns_base.append("tags_as_string")
        for i in range(len(alerts)):
            tags = ""
            try:
                tags = alerts[i]["resource"]["data"]["tags"]
                tags = str(tags)


            except:
                pass

            alerts[i]["tags_as_string"] = tags

    df = pd.json_normalize(alerts)
    if (not df.empty and basic):
        df = df[columns_base]

    if (not df.empty and policyComplianceStandard is not None):
        policies = list_policies(prismaSession=prismasession, complianceStandard=policyComplianceStandard)
        policies = policies[['policyId', 'name', 'description', 'severity', 'recommendation', 'compliance-standard']]
        df = pd.merge(df, policies, how='left', left_on=['policy.policyId'], right_on=['policyId'])
    return df


def login(apiUrl: str = None, access_key_id: str = None, access_key_pass: str = None):
    """
    Login en la api de prisma para empezar a hacer querys.

    :param apiUrl: Cluster donde se encuentra Prisma
    :param access_key_id: Access Key ID
    :param access_key_pass: Access Key Pass
    :return: Objeto con los datos de sesión (url, token)
    """

    ps = prisma.prisma_login(apiUrl, access_key_id, access_key_pass)
    return ps


def query(prismasession: prisma.PrismaSession, query: str = None, timerange=None, return_json=False):
    """
    Hace una query de tipo config search y devuelve resultados.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param query: query a realizar. Debe de ser compatible con las querys de la GUI de Prisma.
    :param timerange: rango de fechas, opcional
    :param return_json: devuelve en formato JSON
    :return: Resultados de la query
    """
    result = []
    res = prisma.config_search(prismasession, query, timerange)
    result += res['items']

    print("Query with pages: " + str(len(result)) + " for now.", end='\r')

    while res['totalRows'] > 0:
        res = prisma.config_search(prismasession, query, timerange, res['nextPageToken'])
        result += res['items']
        print("Query with pages: " + str(len(result)) + " for now.", end='\r')

    if (return_json):
        return result
    else:
        return pd.json_normalize(result)


def disable_accounts(prismasession, account_ids):
    for aid in account_ids:
        res = prisma.disable_account(prismasession, aid)
        print(res)


def enable_accounts(prismasession, account_ids):
    for aid in account_ids:
        res = prisma.enable_account(prismasession, aid)
        print(res)


def alert_cli_remediation(prismasession, alert_id):
    cliRemediation = prisma.alert_cli_remediation(prismasession, alert_id)
    return cliRemediation


def app_security_errors(prismasession):
    """
    Obtiene un listado de errores de IaC, junto con un sumario.
    """

    # summary = prisma.count_iac_errors_by_category(prismasession)
    errors = prisma.code_issues_branch_scans(prismasession)
    summary = {}
    for error in errors:
        if 'codeCategory' in error:
            if error['codeCategory'] not in summary.keys():
                summary[error['codeCategory']] = 0
            summary[error['codeCategory']] += 1

    result = {'summary': summary, 'errors': errors}
    return result
