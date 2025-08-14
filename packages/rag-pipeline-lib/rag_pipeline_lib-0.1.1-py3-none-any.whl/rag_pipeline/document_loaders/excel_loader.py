# rag_pipeline/document_loaders/excel_loader.py
import openpyxl
from typing import List, Dict, Any
from pathlib import Path

class ExcelLoader:
    """Document loader for Excel files (.xlsx)"""

    @staticmethod
    def load_folder(folder_path: str, extensions: List[str] = ['.xlsx']) -> List[Dict[str, Any]]:
        docs: List[Dict[str, Any]] = []
        for ext in extensions:
            for path in Path(folder_path).rglob(f"*{ext}"):
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    text_list = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        text_list.append(f"--- {sheet} ---")
                        for row in ws.iter_rows(values_only=True):
                            row_str = ', '.join([str(cell) if cell is not None else '' for cell in row])
                            text_list.append(row_str)
                    text = '\n'.join(text_list)
                    docs.append({
                        'text': text,
                        'metadata': {
                            'source': str(path),
                            'filename': path.name
                        }
                    })
                except Exception as e:
                    print(f"Error loading Excel {path}: {e}")
        return docs
