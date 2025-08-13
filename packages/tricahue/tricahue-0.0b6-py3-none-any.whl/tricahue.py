from excel2flapjack.main import X2F
import excel2sbol
import sbol2
import requests
import os

import pandas as pd
from openpyxl import load_workbook
import numpy as np
import re

class XDC:

    """XDC class to upload excel file to SynBioHub and Flapjack.

    ...

    Attributes
    ----------
    input_excel_path : str
        path to the input excel file
    fj_url : str
        URL of the Flapjack instance
    fj_user : str
        username of the Flapjack instance
    fj_pass : str
        password of the Flapjack instance
    sbh_url : str
        URL of the SynBioHub instance
    sbh_user : str
        username of the SynBioHub instance
    sbh_pass : str
        password of the SynBioHub instance
    sbh_collection : str
        collection to upload the SBOL file to
    sbh_collection_description : str
        description of the collection
    sbh_overwrite : bool
        whether to overwrite the SBOL file if it already exists
    fj_overwrite : bool
        whether to overwrite the Flapjack project if it already exists  
    fj_token : str
        token to authenticate with Flapjack
    sbh_token : str
        token to authenticate with SynBioHub
    status : str
        status of the process
    
    Methods
    -------
    initialize()
        Initializes the X2F object
    log_in_fj()
        Logs into Flapjack
    log_in_sbh()
        Logs into SynBioHub
    convert_to_sbol()
        Converts the input excel file to SBOL format
    upload_to_fj()
        Uploads the SBOL file to Flapjack
    upload_to_sbh()
        Uploads the SBOL file to SynBioHub
    run()
        Runs the entire process
    """
    def __init__(self, input_excel_path, fj_url, fj_user, fj_pass, sbh_url, sbh_user, sbh_pass, sbh_collection, sbh_collection_description, sbh_overwrite, fj_overwrite, fj_token, sbh_token, homespace):
        self.input_excel_path = input_excel_path
        self.fj_url = fj_url
        self.fj_user = fj_user
        self.fj_pass = fj_pass
        self.sbh_url = sbh_url
        self.sbh_user = sbh_user
        self.sbh_pass = sbh_pass
        self.sbh_collection = sbh_collection
        self.sbh_collection_description = sbh_collection_description
        self.sbh_overwrite = sbh_overwrite
        self.fj_overwrite = fj_overwrite
        self.fj_token = fj_token
        self.sbh_token = sbh_token
        self.input_excel = pd.ExcelFile(self.input_excel_path)
        self.x2f = None
        self.sbol_doc = None
        self.sbol_fj_doc = None
        self.sbol_graph_uri = None
        self.file_path_out = f'{sbh_collection}_converted_SBOL.xml'
        self.file_path_out2 = f'{sbh_collection}_SBOL_Fj_doc.xml'
        self.homespace = homespace
        self.sbol_hash_map = {}

    def initialize(self):
        self.x2f = X2F(excel_path=self.input_excel_path,
                    fj_url=self.fj_url, 
                    overwrite=self.fj_overwrite)
        if self.sbh_collection_description is None:
            self.sbh_collection_description = 'Collection of SBOL files uploaded from Tricahue'
        if self.sbol_doc is None:
            self.sbol_doc = sbol2.Document()
        if self.sbol_fj_doc is None:
            self.sbol_fj_doc = sbol2.Document()
        
    def log_in_fj(self):
        self.x2f = X2F(excel_path=self.input_excel_path, 
                    fj_url=self.fj_url, 
                    overwrite=self.fj_overwrite)
        
        if self.fj_token:
            self.x2f.fj.log_in_token(username=self.fj_user, access_token=None, refresh_token=self.fj_token)
            self.x2f.fj.refresh()

        elif self.fj_user and self.fj_pass:
            self.x2f.fj.log_in(username=self.fj_user, password=self.fj_pass)
            self.fj_token = self.x2f.fj.refresh_token
        
        else:
            print('Unable to authenticate into Flapjack')
            #TODO check token validity
        

    def log_in_sbh(self):
        # SBH Login
        if self.sbh_token is None:
            response = requests.post(
                f'{self.sbh_url}/login',
                headers={'Accept': 'text/plain'},
                data={
                    'email': self.sbh_user,
                    'password' : self.sbh_pass,
                    }
            )
            self.sbh_token = response.text
        else:
            response = requests.post(
                f'{self.sbh_url}/login',
                headers={'Accept': 'text/plain', 'X-Authorization': self.sbh_token}
            )

    def convert_to_sbol(self, sbol_version=2):
        excel2sbol.converter(file_path_in = self.input_excel_path, 
                file_path_out = self.file_path_out, homespace=self.homespace, sbol_version=sbol_version)
        doc = sbol2.Document()
        doc.read(self.file_path_out)
        self.sbol_doc = doc        

    def generate_sbol_hash_map(self):
        # Pull graph uri from synbiohub
        response = requests.get(
            f'{self.sbh_url}/profile',
            headers={
                'Accept': 'text/plain',
                'X-authorization': self.sbh_token
                }
        )
        self.sbol_graph_uri = response.json()['graphUri']
        sbol_collec_url = f'{self.sbol_graph_uri}/{self.sbh_collection}'

        # create hashmap of flapjack id to sbol uri
        self.sbol_hash_map = {}
        for tl in self.sbol_doc:
            #if 'https://flapjack.rudge-lab.org/ID' in tl.properties:
            sbol_uri = tl.properties['http://sbols.org/v2#persistentIdentity'][0]
            sbol_uri = sbol_uri.replace(self.homespace, sbol_collec_url)
            sbol_uri = f'{sbol_uri}/1'

            sbol_name = str(tl.properties['http://sbols.org/v2#displayId'][0])
            self.sbol_hash_map[sbol_name] = sbol_uri


    def upload_to_fj(self, header_rows=3):
        self.x2f.sbol_hash_map = self.sbol_hash_map
        self.x2f.generate_sheets_to_object_mapping()
        self.x2f.index_skiprows = header_rows
        # self.x2f.create_df()
        # change to upload_object_in_sheets
        self.x2f.upload_all() 


    def upload_to_sbh(self):
        # Add flapjack annotations to the SBOL
        doc = sbol2.Document()
        doc.read(self.file_path_out)
        for tl in self.sbol_doc:
            id = str(tl).split('/')[-2]
            if id in self.sbol_hash_map:
                setattr(tl, 'Flapjack_ID',
                        sbol2.URIProperty(tl,
                        f'https://flapjack.rudge-lab.org/ID',
                            '0', '1', [], initial_value=f'http://wwww.{self.fj_url}/{self.sbol_hash_map[id]}'))
        #doc = sbol2.Document()
        doc.write(self.file_path_out2)

        if self.sbh_overwrite:
            sbh_overwrite = '1'
        else:
            sbh_overwrite = '0'
        # SBH file upload
        response = requests.post(
            f'{self.sbh_url}/submit',
            headers={
                'Accept': 'text/plain',
                'X-authorization': self.sbh_token
            },
            files={
            'files': open(self.file_path_out2,'rb'),
            },
            data={
                'id': self.sbh_collection,
                'version' : '1',
                'name' : self.sbh_collection,
                'description' : self.sbh_collection_description, #TODO
                'overwrite_merge' : self.sbh_overwrite
            },

        )

        if response.text == "Submission id and version already in use":
            print('not submitted')
            raise AttributeError(f'The collection ({self.sbh_collection}) could not be submitted to synbiohub as the collection already exists and overite is not on.')
        # if response.text == "Successfully uploaded":
        #      success = True
        #self.status = "Uploaded to SynBioHub"
        return f'{self.sbol_graph_uri}/{self.sbh_collection}/{self.sbh_collection}_collection/1'

        

    def run(self):
        self.initialize()
        self.log_in_fj()
        self.log_in_sbh()
        self.convert_to_sbol()
        self.generate_sbol_hash_map()
        self.upload_to_fj()
        self.upload_to_sbh()


class XDE:

    """XDE (Experimental Data Extractor) class to extract experimental data from
    plate reader excel output and writes it in an XDC template.

    ...

    Attributes
    ----------


    Methods
    -------
    getFileNameFromString(string)
        Extracts the file name from a string
    generateSampleData(file_list,sheet_to_read_from,time_col_name,data_cols_offset)
        Generates sample data from the input excel files
    getNumRows(dataframe,starting_row_idx,starting_col_idx)
        Gets the number of rows for the data
    buildFinalDF(file_list,sample_data_list,time_col_name,data_cols_offset,num_rows_btwn_data,sheet_to_read_from)
        Builds the final dataframe
    writeToMeasurements(XDC_file_name,final_dataframe)
        Writes the final dataframe to the measurements sheet
    extractData(file_list,sheet_to_read_from,time_col_name,data_cols_offset,num_rows_btwn_data)
        Full run; extracts data from the input excel files and writes it to the XDC sheet

    """
    def getFileNameFromString(self, string):
        pattern = '[\w-]+?(?=\.)'
        # searching the pattern
        result = re.search(pattern, string)
    
        return result.group()

    def generateSampleData(self, file_list, sheet_to_read_from,time_col_name, data_cols_offset=0): 
        num_assays = len(file_list) - 1
        file_name_list = []

        for i in range(num_assays):
            file_name_list.append(self.getFileNameFromString(file_list[i + 1]))

        #final products
        result = pd.DataFrame()
        sample_data_list = []

        #components:result
        assay_id = []
        column = []
        row = []
        sample_id = []

        #componenets:sample_data_list
        columnID = []
        assay_num = []

        #processing:main
        current_sample_id = 1

        for i in range(num_assays):

            current_num_assay = i + 1

            #locating instances of time_col_name
            raw_df = pd.read_excel(file_list[i+1],sheet_to_read_from)
            rows, cols = np.where(raw_df == time_col_name)
            time_col_locations = list(zip(rows, cols))
            num_rows = self.getNumRows(raw_df,rows[0],cols[0])
            
            #extracting signal 1 data to check for blank columns
            start_row = time_col_locations[0][0] + 1
            start_col = data_cols_offset
            num_cols = 96
            working_df = raw_df.iloc[start_row:start_row + num_rows, start_col:start_col + num_cols] #maybe subtract 1 from num rows

            # Check for completely blank (all NaN) columns using numpy
            is_blank = working_df.isna().all().to_numpy()

            # Get the indices of non-blank columns
            data_col_IDX = np.where(~is_blank)[0]
            
            #add to lists
            for j in range(len(data_col_IDX)):
                #result
                assay_id.append(file_name_list[i])
                column.append(data_col_IDX[j] % 12 + 1) #IDX % 12 + 1
                row.append((data_col_IDX[j]//12) + 1)   #IDX // 12 + 1
                sample_id.append(f"Sample{current_sample_id}")
                current_sample_id += 1

                #sample_data_list
                columnID.append(data_col_IDX[j])
                assay_num.append(current_num_assay)

        #assembly:result
        result.insert(0, "Assay ID", assay_id)
        result.insert(0, "Column", column)
        result.insert(0, "Row", row)
        result.insert(0, "Sample ID", sample_id)
        
        #assembly:sample_data_list
        for i in range(len(result)):
            temp_tuple = (sample_id[i],columnID[i],assay_num[i])
            sample_data_list.append(temp_tuple)
        
        with pd.ExcelWriter(file_list[0], mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            result.to_excel(writer,'Sample',startrow=3,index=False)

        return sample_data_list

    def getNumRows(self, dataframe, starting_row_idx, starting_col_idx):
        num_rows = 0
        counter = 1
        time_col_value = dataframe.iloc[starting_row_idx, starting_col_idx]

        while True:
            current_cell = dataframe.iloc[starting_row_idx + counter, starting_col_idx]
            if pd.isna(current_cell) or current_cell == time_col_value:
                break
            if(len(dataframe) <= counter + starting_row_idx + 1): #edge case for if there is only one signal, IDK why i have to add a +1
                num_rows += 1
                break

            counter += 1
            num_rows += 1
            
        return num_rows 

    def buildFinalDF(self, file_list, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data, sheet_to_read_from):
        print(file_list)
        output = pd.DataFrame() 
        time_col_locations = []
        num_rows_per_assay = []
        dataframe_list = []
        num_assays = len(file_list) - 1

        for i in range(num_assays):
            raw_df = pd.read_excel(file_list[i+1],sheet_to_read_from)
            rows, cols = np.where(raw_df == time_col_name)
            temp = list(zip(rows, cols))
            num_rows_per_assay.append(self.getNumRows(raw_df,rows[0],cols[0]))
            time_col_locations.append(temp)
            dataframe_list.append(pd.read_excel(file_list[i + 1],sheet_to_read_from))
            
        for i in range(len(sample_data_list)):  #initilizing information about the current sample and its results
            rows_to_be_read = []
            current_sample_id = str(sample_data_list[i][0])               
            current_col = sample_data_list[i][1]
            current_assay = sample_data_list[i][2]
            current_first_row = time_col_locations[current_assay - 1][0][0] + 1
            current_time_col = time_col_locations[current_assay - 1][0][1]
            current_num_rows = num_rows_per_assay[current_assay - 1]
            current_num_signals = len(time_col_locations[current_assay - 1])

            for j in range(current_num_signals): 
                rows_to_be_read.extend(list(range(current_first_row + ((current_num_rows + num_rows_btwn_data + 1)* j), current_first_row + current_num_rows + ((current_num_rows + num_rows_btwn_data + 1)* j))))
            working_df = dataframe_list[current_assay - 1].iloc[rows_to_be_read,[current_time_col,current_col + data_cols_offset]].copy() # at this point it will be the time col and current col for both signals
            working_df.columns = ["Time", "Value"]
            #add signal label
            signal_id = []
            for k in range(current_num_signals):
                signal_id.extend([f"Signal{k + 1}"] * current_num_rows)
            working_df.insert(0, "Signal ID", signal_id)

            #add sample label
            sample_id = [current_sample_id] * len(working_df)
            working_df.insert(0, "Sample ID", sample_id)

            #concat working_df and output
            output = pd.concat([output, working_df], ignore_index=True)

        #add measurement
        measurement_id = []
        for i in range(len(output)):
            measurement_id.append(f"Measurement{i}")
        output.insert(0, "Measurement ID", measurement_id)

        return output

    def buildFinalDFCSV(self, file_list, sample_data_list, time_col_name, data_cols_offset, num_rows_btwn_data):
        output = pd.DataFrame() 
        time_col_locations = []
        num_rows_per_assay = []
        dataframe_list = []
        num_assays = len(file_list) - 1

        for i in range(num_assays):
            raw_df = pd.read_csv(file_list[i+1])
            rows, cols = np.where(raw_df == time_col_name)
            temp = list(zip(rows, cols))
            num_rows_per_assay.append(self.getNumRows(raw_df,rows[0],cols[0]))
            time_col_locations.append(temp)
            dataframe_list.append(pd.read_csv(file_list[i + 1]))
            
        for i in range(len(sample_data_list)):  #initilizing information about the current sample and its results
            rows_to_be_read = []
            current_sample_id = sample_data_list[i][0]               
            current_col = sample_data_list[i][1]
            current_assay = sample_data_list[i][2]
            current_first_row = time_col_locations[current_assay - 1][0][0] + 1
            current_time_col = time_col_locations[current_assay - 1][1][1]
            current_num_rows = num_rows_per_assay[current_assay - 1]
            current_num_signals = len(time_col_locations[current_assay - 1])

            for j in range(current_num_signals): 
                rows_to_be_read.extend(list(range(current_first_row + ((current_num_rows + num_rows_btwn_data + 1)* j), current_first_row + current_num_rows + ((current_num_rows + num_rows_btwn_data + 1)* j))))
            
            working_df = dataframe_list[current_assay - 1].iloc[rows_to_be_read,[current_time_col,current_col + data_cols_offset]].copy() # at this point it will be the time col and current col for both signals
            working_df.columns = ["Time", "Value"]

            #add signal label
            signal_id = []
            for k in range(current_num_signals):
                signal_id.extend([f"Signal{k + 1}"] * current_num_rows)
            working_df.insert(0, "Signal ID", signal_id)

            #add sample label
            sample_id = [current_sample_id] * len(working_df)
            working_df.insert(0, "Sample ID", sample_id)

            #concat working_df and output
            output = pd.concat([output, working_df], ignore_index=True)

        #add measurement
        measurement_id = []
        for i in range(len(output)):
            measurement_id.append(f"Measurement{i}")
        output.insert(0, "Measurement ID", measurement_id)

        return output


    def writeToMeasurements(self, XDC_file_name, final_dataframe):
        book = load_workbook(XDC_file_name)
        sheet = book['Measurement']

        # Clear the existing data in the 'Measurement' sheet
        sheet.delete_rows(1, sheet.max_row)

        # Write three blank rows before writing the data
        for _ in range(3):
            sheet.append([''] * 5)

        # Write the headers
        sheet.append(['Measurement ID', 'Sample ID', 'Signal ID', 'Time', 'Value'])

        # Write the data
        for row in final_dataframe.itertuples(index=False):
            sheet.append(list(row))

        book.save(XDC_file_name)
        book.close()

        return
    
    def extractData(self, file_list, sheet_to_read_from, time_col_name, data_cols_offset, num_rows_btwn_data=0):
        """
        Full run; extracts data from the input excel files and writes it to the XDC sheet.
        """
        sample_list = self.generateSampleData(file_list, sheet_to_read_from, time_col_name, data_cols_offset)
        output_df = self.buildFinalDF(file_list, sample_list, time_col_name, data_cols_offset, num_rows_btwn_data, sheet_to_read_from)
        self.writeToMeasurements(file_list[0], output_df)